"""Gera Excel formatado a partir do CSV king-inativos-com-produtos.csv"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime

CSV = r"C:\Users\teccr\gestaocrosby\.tmp-test\king-inativos-com-produtos.csv"
XLSX = r"C:\Users\teccr\gestaocrosby\.tmp-test\king-inativos-com-produtos.xlsx"
PROMO = r"C:\Users\teccr\gestaocrosby\.tmp-test\promo-varejo.json"

df = pd.read_csv(CSV, dtype={
    'person_code': str, 'cpf_cnpj': str, 'filial_vendedora': str, 'vendedor_codigo': str
})
df['ultima_compra'] = pd.to_datetime(df['ultima_compra'], errors='coerce')
df['ultima_compra_king'] = pd.to_datetime(df['ultima_compra_king'], errors='coerce')
total_bruto = len(df)
# Remove BAZAR/SALDO (clientes de saldão/defeito não são alvo de campanha King)
df = df[df['categoria'] != 'BAZAR/SALDO'].reset_index(drop=True)
print(f"Removidos {total_bruto - len(df)} BAZAR/SALDO. Restam {len(df)} linhas.")

wb = Workbook()

# ==== Sheet 1: Inativos King ====
ws = wb.active
ws.title = "Inativos King"

# Header
HEADER_FILL = PatternFill("solid", start_color="000638")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
BODY_FONT = Font(name="Calibri", size=10)

cols = list(df.columns)
for j, col in enumerate(cols, 1):
    c = ws.cell(row=1, column=j, value=col)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER

# Data
for i, row in enumerate(df.itertuples(index=False), 2):
    for j, val in enumerate(row, 1):
        c = ws.cell(row=i, column=j)
        if pd.isna(val):
            c.value = ""
        elif isinstance(val, pd.Timestamp):
            c.value = val.date()
            c.number_format = "yyyy-mm-dd"
        else:
            c.value = val
        c.font = BODY_FONT
        c.alignment = LEFT

# Column widths
widths = {
    'person_code': 11, 'person_name': 38, 'categoria': 18,
    'operation_dominante': 40, 'cpf_cnpj': 16, 'tipo_pessoa': 6,
    'fantasy_name': 30, 'uf': 5, 'telefone': 16, 'email': 32,
    'filial_vendedora': 10, 'vendedor_codigo': 10, 'vendedor_nome': 30,
    'ultima_compra': 13, 'ultima_compra_king': 14, 'dias_inativo': 11,
    'qtd_skus_king': 10, 'produtos_king_comprados': 80,
}
for j, col in enumerate(cols, 1):
    ws.column_dimensions[get_column_letter(j)].width = widths.get(col, 15)

# Freeze + AutoFilter
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(df) + 1}"

# Conditional formatting on dias_inativo
dias_col = cols.index("dias_inativo") + 1
dias_letter = get_column_letter(dias_col)
data_range = f"{dias_letter}2:{dias_letter}{len(df) + 1}"
green_fill = PatternFill("solid", start_color="C6EFCE")
yellow_fill = PatternFill("solid", start_color="FFEB9C")
red_fill = PatternFill("solid", start_color="FFC7CE")
ws.conditional_formatting.add(data_range, CellIsRule(operator="lessThan", formula=["240"], fill=green_fill))
ws.conditional_formatting.add(data_range, CellIsRule(operator="between", formula=["240", "365"], fill=yellow_fill))
ws.conditional_formatting.add(data_range, CellIsRule(operator="greaterThan", formula=["365"], fill=red_fill))

# Row height
ws.row_dimensions[1].height = 26

# ==== Sheet 2: Resumo ====
ws2 = wb.create_sheet("Resumo")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="000638")
SUB_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
SUB_FILL = PatternFill("solid", start_color="000638")
TOTAL_FONT = Font(name="Calibri", size=11, bold=True)

ws2['A1'] = "📊 King Inativos 6+ meses — Resumo"
ws2['A1'].font = TITLE_FONT
ws2.merge_cells('A1:D1')

# Bloco: Categoria
r = 3
ws2.cell(row=r, column=1, value="Por categoria").font = SUB_FONT
ws2.cell(row=r, column=1).fill = SUB_FILL
ws2.cell(row=r, column=2, value="Clientes").font = SUB_FONT
ws2.cell(row=r, column=2).fill = SUB_FILL
ws2.cell(row=r, column=2).alignment = CENTER
r += 1
cat_counts = df['categoria'].value_counts()
cat_start = r
for cat, cnt in cat_counts.items():
    ws2.cell(row=r, column=1, value=cat)
    ws2.cell(row=r, column=2, value=cnt).alignment = CENTER
    r += 1
ws2.cell(row=r, column=1, value="TOTAL").font = TOTAL_FONT
total_cell = ws2.cell(row=r, column=2, value=f"=SUM(B{cat_start}:B{r-1})")
total_cell.font = TOTAL_FONT
total_cell.alignment = CENTER
# Cache valor calculado pra evitar dependência de recálculo LibreOffice
total_cell.value = int(cat_counts.sum())

# Bloco: Filial
r += 3
ws2.cell(row=r, column=1, value="TOP filiais que venderam King").font = SUB_FONT
ws2.cell(row=r, column=1).fill = SUB_FILL
ws2.cell(row=r, column=2, value="Clientes").font = SUB_FONT
ws2.cell(row=r, column=2).fill = SUB_FILL
ws2.cell(row=r, column=2).alignment = CENTER
r += 1
filial_counts = df['filial_vendedora'].value_counts().head(10)
fil_start = r
for fil, cnt in filial_counts.items():
    ws2.cell(row=r, column=1, value=f"Filial {fil}")
    ws2.cell(row=r, column=2, value=int(cnt)).alignment = CENTER
    r += 1

# Bloco: UF
r += 3
ws2.cell(row=r, column=1, value="Por UF").font = SUB_FONT
ws2.cell(row=r, column=1).fill = SUB_FILL
ws2.cell(row=r, column=2, value="Clientes").font = SUB_FONT
ws2.cell(row=r, column=2).fill = SUB_FILL
ws2.cell(row=r, column=2).alignment = CENTER
r += 1
uf_counts = df['uf'].fillna('(sem UF)').replace('', '(sem UF)').value_counts().head(15)
for uf, cnt in uf_counts.items():
    ws2.cell(row=r, column=1, value=uf)
    ws2.cell(row=r, column=2, value=int(cnt)).alignment = CENTER
    r += 1

# Bloco: TOP SKUs (calculado das colunas produtos_king_comprados)
sku_counter = {}
sku_name = {}
for prods in df['produtos_king_comprados'].fillna('').astype(str):
    if not prods:
        continue
    for item in prods.split(' | '):
        parts = item.split(' ', 1)
        if not parts or not parts[0].isdigit():
            continue
        sku = parts[0]
        sku_counter[sku] = sku_counter.get(sku, 0) + 1
        if sku not in sku_name and len(parts) > 1:
            # Remove a parte do preço pós "—"
            name = parts[1].split(' — ')[0].rsplit(' ×', 1)[0]
            sku_name[sku] = name

# Carrega preços do JSON promo
import json
with open(PROMO, 'r', encoding='utf-8') as f:
    promo = json.load(f)
promo_items = promo.get('items', []) if isinstance(promo, dict) else promo
promo_by_sku = {str(p['sku']): p for p in promo_items}

top_skus = sorted(sku_counter.items(), key=lambda x: -x[1])[:15]

r += 3
ws2.cell(row=r, column=1, value="TOP 15 SKUs King mais comprados pelos inativos").font = SUB_FONT
ws2.cell(row=r, column=1).fill = SUB_FILL
ws2.cell(row=r, column=2, value="Clientes").font = SUB_FONT
ws2.cell(row=r, column=2).fill = SUB_FILL
ws2.cell(row=r, column=2).alignment = CENTER
ws2.cell(row=r, column=3, value="Preço promo").font = SUB_FONT
ws2.cell(row=r, column=3).fill = SUB_FILL
ws2.cell(row=r, column=3).alignment = CENTER
ws2.cell(row=r, column=4, value="Desconto").font = SUB_FONT
ws2.cell(row=r, column=4).fill = SUB_FILL
ws2.cell(row=r, column=4).alignment = CENTER
r += 1
for sku, cnt in top_skus:
    p = promo_by_sku.get(sku, {})
    nome = sku_name.get(sku, p.get('descricao', '?'))
    ws2.cell(row=r, column=1, value=f"{sku}  {nome}")
    ws2.cell(row=r, column=2, value=cnt).alignment = CENTER
    if p.get('preco_promo') is not None:
        ws2.cell(row=r, column=3, value=float(p['preco_promo'])).number_format = 'R$ #,##0.00'
    if p.get('desconto_pct') is not None:
        ws2.cell(row=r, column=4, value=float(p['desconto_pct']) / 100).number_format = '0%'
    r += 1

# Larguras Resumo
ws2.column_dimensions['A'].width = 55
ws2.column_dimensions['B'].width = 12
ws2.column_dimensions['C'].width = 14
ws2.column_dimensions['D'].width = 12

# Header row height
ws2.row_dimensions[1].height = 22

wb.save(XLSX)
print(f"OK salvo em {XLSX}")
