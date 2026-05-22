"""Gera Excel formatado para king-multimarcas-tudo.csv (146 clientes, sem filtro de data)"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
import json

CSV = r"C:\Users\teccr\gestaocrosby\.tmp-test\king-multimarcas-tudo.csv"
CSV_DET = r"C:\Users\teccr\gestaocrosby\.tmp-test\king-multimarcas-detalhado.csv"
XLSX = r"C:\Users\teccr\gestaocrosby\.tmp-test\king-multimarcas-tudo.xlsx"
PROMO = r"C:\Users\teccr\gestaocrosby\.tmp-test\promo-varejo.json"

df = pd.read_csv(CSV, dtype={
    'person_code': str, 'cpf_cnpj': str, 'filial_vendedora': str,
    'vendedor_codigo': str, 'telefone': str
})
df['primeira_compra_king'] = pd.to_datetime(df['primeira_compra_king'], errors='coerce')
df['ultima_compra_king'] = pd.to_datetime(df['ultima_compra_king'], errors='coerce')
df['valor_total_king'] = pd.to_numeric(df['valor_total_king'], errors='coerce')
print(f"Lidas {len(df)} linhas")

wb = Workbook()
ws = wb.active
ws.title = "Multimarcas King"

HEADER_FILL = PatternFill("solid", start_color="000638")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
BODY_FONT = Font(name="Calibri", size=10)

cols = list(df.columns)
for j, col in enumerate(cols, 1):
    c = ws.cell(row=1, column=j, value=col)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER

for i, row in enumerate(df.itertuples(index=False), 2):
    for j, val in enumerate(row, 1):
        c = ws.cell(row=i, column=j)
        col_name = cols[j-1]
        if pd.isna(val):
            c.value = ""
        elif isinstance(val, pd.Timestamp):
            c.value = val.date()
            c.number_format = "yyyy-mm-dd"
        elif col_name == 'valor_total_king':
            c.value = float(val) if pd.notna(val) else 0
            c.number_format = 'R$ #,##0.00'
        else:
            c.value = val
        c.font = BODY_FONT
        c.alignment = LEFT

widths = {
    'person_code': 11, 'person_name': 38, 'cpf_cnpj': 16, 'tipo_pessoa': 6,
    'fantasy_name': 30, 'uf': 5, 'telefone': 16, 'email': 32,
    'operation_dominante': 40, 'filial_vendedora': 10, 'vendedor_codigo': 10,
    'vendedor_nome': 28, 'primeira_compra_king': 14, 'ultima_compra_king': 14,
    'dias_inativo': 11, 'qtd_nfs_king': 11, 'qtd_skus_distintos': 11,
    'qtd_pecas_king': 12, 'valor_total_king': 16, 'produtos_king_comprados': 80,
}
for j, col in enumerate(cols, 1):
    ws.column_dimensions[get_column_letter(j)].width = widths.get(col, 15)

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(df) + 1}"
ws.row_dimensions[1].height = 26

# Cond. fmt dias_inativo
dias_col = cols.index("dias_inativo") + 1
dias_letter = get_column_letter(dias_col)
data_range = f"{dias_letter}2:{dias_letter}{len(df) + 1}"
green = PatternFill("solid", start_color="C6EFCE")
yellow = PatternFill("solid", start_color="FFEB9C")
red = PatternFill("solid", start_color="FFC7CE")
ws.conditional_formatting.add(data_range, CellIsRule(operator="lessThan", formula=["180"], fill=green))
ws.conditional_formatting.add(data_range, CellIsRule(operator="between", formula=["180", "365"], fill=yellow))
ws.conditional_formatting.add(data_range, CellIsRule(operator="greaterThan", formula=["365"], fill=red))

# ==== Resumo ====
ws2 = wb.create_sheet("Resumo")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="000638")
SUB_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
SUB_FILL = PatternFill("solid", start_color="000638")
TOTAL_FONT = Font(name="Calibri", size=11, bold=True)

ws2['A1'] = "King × Multimarcas — Histórico completo"
ws2['A1'].font = TITLE_FONT
ws2.merge_cells('A1:D1')

r = 3
ws2.cell(r, 1, "INDICADOR").font = SUB_FONT; ws2.cell(r, 1).fill = SUB_FILL
ws2.cell(r, 2, "VALOR").font = SUB_FONT; ws2.cell(r, 2).fill = SUB_FILL; ws2.cell(r, 2).alignment = CENTER
r += 1
total_valor = float(df['valor_total_king'].sum())
total_pecas = int(df['qtd_pecas_king'].sum())
total_nfs = int(df['qtd_nfs_king'].sum())
ativos = int((df['dias_inativo'] <= 180).sum())
inativos_6_12 = int(((df['dias_inativo'] > 180) & (df['dias_inativo'] <= 365)).sum())
inativos_1ano = int((df['dias_inativo'] > 365).sum())

linhas_resumo = [
    ("Clientes multimarcas únicos", len(df), None),
    ("NFs King-multimarcas (histórico)", total_nfs, None),
    ("Peças King vendidas (acumulado)", total_pecas, None),
    ("Valor total King (líquido)", total_valor, 'R$ #,##0.00'),
    ("Ativos (King últimos 180d)", ativos, None),
    ("Inativos 6-12 meses", inativos_6_12, None),
    ("Inativos 12+ meses", inativos_1ano, None),
]
for label, val, fmt in linhas_resumo:
    ws2.cell(r, 1, label)
    cell = ws2.cell(r, 2, val)
    cell.alignment = CENTER
    if fmt: cell.number_format = fmt
    r += 1

# Top filial
r += 2
ws2.cell(r, 1, "TOP filiais que venderam King-MM").font = SUB_FONT; ws2.cell(r, 1).fill = SUB_FILL
ws2.cell(r, 2, "Clientes").font = SUB_FONT; ws2.cell(r, 2).fill = SUB_FILL; ws2.cell(r, 2).alignment = CENTER
ws2.cell(r, 3, "Valor total").font = SUB_FONT; ws2.cell(r, 3).fill = SUB_FILL; ws2.cell(r, 3).alignment = CENTER
r += 1
filial_grp = df.groupby('filial_vendedora').agg(
    clientes=('person_code', 'count'),
    valor=('valor_total_king', 'sum'),
).sort_values('valor', ascending=False).head(10)
for fil, g in filial_grp.iterrows():
    ws2.cell(r, 1, f"Filial {fil}")
    ws2.cell(r, 2, int(g['clientes'])).alignment = CENTER
    ws2.cell(r, 3, float(g['valor'])).number_format = 'R$ #,##0.00'
    r += 1

# Por UF
r += 2
ws2.cell(r, 1, "Por UF").font = SUB_FONT; ws2.cell(r, 1).fill = SUB_FILL
ws2.cell(r, 2, "Clientes").font = SUB_FONT; ws2.cell(r, 2).fill = SUB_FILL; ws2.cell(r, 2).alignment = CENTER
ws2.cell(r, 3, "Valor total").font = SUB_FONT; ws2.cell(r, 3).fill = SUB_FILL; ws2.cell(r, 3).alignment = CENTER
r += 1
uf_grp = df.assign(uf_n=df['uf'].fillna('(sem UF)').replace('', '(sem UF)')).groupby('uf_n').agg(
    clientes=('person_code', 'count'),
    valor=('valor_total_king', 'sum'),
).sort_values('valor', ascending=False)
for uf, g in uf_grp.iterrows():
    ws2.cell(r, 1, uf)
    ws2.cell(r, 2, int(g['clientes'])).alignment = CENTER
    ws2.cell(r, 3, float(g['valor'])).number_format = 'R$ #,##0.00'
    r += 1

# TOP SKUs
with open(PROMO, 'r', encoding='utf-8') as f:
    promo = json.load(f)
promo_items = promo.get('items', []) if isinstance(promo, dict) else promo
promo_by_sku = {str(p['sku']): p for p in promo_items}

sku_count = {}
sku_pecas = {}
for prods in df['produtos_king_comprados'].fillna('').astype(str):
    seen = set()
    for item in prods.split(' | '):
        if not item: continue
        parts = item.split(' ')
        if not parts[0].isdigit(): continue
        sku = parts[0]
        # extrai qtd ×N
        qtd = 0
        for p in parts:
            if p.startswith('×'):
                try: qtd = int(p[1:])
                except: pass
        if sku not in seen:
            sku_count[sku] = sku_count.get(sku, 0) + 1
            seen.add(sku)
        sku_pecas[sku] = sku_pecas.get(sku, 0) + qtd

top_skus = sorted(sku_count.items(), key=lambda x: -x[1])[:15]
r += 2
ws2.cell(r, 1, "TOP 15 SKUs King mais comprados (multimarcas)").font = SUB_FONT
ws2.cell(r, 1).fill = SUB_FILL
ws2.cell(r, 2, "Clientes").font = SUB_FONT; ws2.cell(r, 2).fill = SUB_FILL; ws2.cell(r, 2).alignment = CENTER
ws2.cell(r, 3, "Peças").font = SUB_FONT; ws2.cell(r, 3).fill = SUB_FILL; ws2.cell(r, 3).alignment = CENTER
ws2.cell(r, 4, "Preço promo").font = SUB_FONT; ws2.cell(r, 4).fill = SUB_FILL; ws2.cell(r, 4).alignment = CENTER
r += 1
for sku, cnt in top_skus:
    p = promo_by_sku.get(sku, {})
    nome = f"{sku}  {p.get('produto', '?')} {p.get('cor', '')} {p.get('tam', '')}".strip()
    ws2.cell(r, 1, nome)
    ws2.cell(r, 2, cnt).alignment = CENTER
    ws2.cell(r, 3, sku_pecas.get(sku, 0)).alignment = CENTER
    if p.get('preco_promo') is not None:
        ws2.cell(r, 4, float(p['preco_promo'])).number_format = 'R$ #,##0.00'
    r += 1

ws2.column_dimensions['A'].width = 60
ws2.column_dimensions['B'].width = 12
ws2.column_dimensions['C'].width = 14
ws2.column_dimensions['D'].width = 14
ws2.row_dimensions[1].height = 22

# ==== Aba 3: Vendas Detalhadas (1 linha por NF×SKU) ====
det = pd.read_csv(CSV_DET, dtype={
    'person_code': str, 'cpf_cnpj': str, 'branch_code': str, 'dealer_code': str,
    'sku': str, 'telefone': str
})
det['issue_date'] = pd.to_datetime(det['issue_date'], errors='coerce')
for col_n in ['preco_unitario', 'preco_promo_atual', 'valor_total_nf_sku', 'desconto_pct', 'qtd']:
    det[col_n] = pd.to_numeric(det[col_n], errors='coerce')
print(f"Vendas detalhadas: {len(det)} linhas")

ws3 = wb.create_sheet("Vendas Detalhadas")
det_cols = list(det.columns)
for j, col in enumerate(det_cols, 1):
    c = ws3.cell(row=1, column=j, value=col)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER

for i, row in enumerate(det.itertuples(index=False), 2):
    for j, val in enumerate(row, 1):
        c = ws3.cell(row=i, column=j)
        col_name = det_cols[j-1]
        if pd.isna(val) or val == '':
            c.value = ""
        elif isinstance(val, pd.Timestamp):
            c.value = val.date()
            c.number_format = "yyyy-mm-dd"
        elif col_name in ('preco_unitario', 'preco_promo_atual', 'valor_total_nf_sku'):
            c.value = float(val) if pd.notna(val) else 0
            c.number_format = 'R$ #,##0.00'
        elif col_name == 'desconto_pct':
            c.value = float(val) / 100 if pd.notna(val) else 0
            c.number_format = '0%'
        else:
            c.value = val
        c.font = BODY_FONT
        c.alignment = LEFT

det_widths = {
    'issue_date': 12, 'person_code': 11, 'person_name': 32, 'cpf_cnpj': 16,
    'fantasy_name': 28, 'telefone': 14, 'operation_name': 35,
    'branch_code': 9, 'dealer_code': 9, 'vendedor_nome': 28,
    'sku': 8, 'produto': 36, 'produto_promo': 32,
    'preco_unitario': 13, 'preco_promo_atual': 14, 'desconto_pct': 11,
    'qtd': 7, 'valor_total_nf_sku': 14,
}
for j, col in enumerate(det_cols, 1):
    ws3.column_dimensions[get_column_letter(j)].width = det_widths.get(col, 14)
ws3.freeze_panes = "A2"
ws3.auto_filter.ref = f"A1:{get_column_letter(len(det_cols))}{len(det) + 1}"
ws3.row_dimensions[1].height = 26

wb.save(XLSX)
print(f"OK salvo: {XLSX}")
